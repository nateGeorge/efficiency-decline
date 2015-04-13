from openpyxl import load_workbook
import re, time, csv, glob, os
import numpy as np
import pylab as plt

def sort_pcbe_tool(rundata):
	newrundata={}
	MC01runs=[]
	MC02runs=[]
	for each in rundata:
		try:
			runnum=int(each[0])
		except TypeError:
			runnum=1
		if runnum>=290:
			runnum=str(runnum)
			newrundata[runnum]={}
			newrundata[runnum]['PCBEtool']=each[1]
			newrundata[runnum]['runlen']=each[2]
			newrundata[runnum]['rundescr']=each[3]
			if each[1]=='MC01':
				MC01runs.append(each)
			else:
				MC02runs.append(each)
	return newrundata, MC01runs, MC02runs

	
effCutoff=9.9	

f='Y:\ProcessFE\DailyWhiteBoard_2015_1Q ACTIVE.xlsx'
wb = load_workbook(filename = f,use_iterators=True, data_only=True)	
ws4 = wb.get_sheet_by_name(name="Runs")


porinclude=[334,376]
porexculde=[299,303,305,307,316,331,332,339,344,361]
porrundata=[]
mr600rundata=[]
SPrundata=[]
otherrundata=[]
allrundata={}
rowcounter=0
for row in ws4.iter_rows():
	if rowcounter==0:
		labels=[row[0].value,row[6].value,row[3].value,row[10].value.encode('utf-8')]
	if row[10].value!=None and row[0].value!=None and rowcounter>0 and isinstance(row[0].value,float):
		added=False
		thedescr=row[10].value.encode('utf-8')#.decode('utf-8')
		#thedescr=thedescr.encode('CP1252')
		runnum=int(row[0].value)
		PCBEtool=row[6].value
		runlength=row[3].value
		allrundata[str(runnum)]=[PCBEtool,runlength]
		if runnum not in porexculde and re.search('POR',thedescr) or runnum in porinclude:
			if not re.search('POR-ish',thedescr):
				porrundata.append((runnum,PCBEtool,runlength,thedescr))
				allrundata[str(runnum)].append('POR')
				added=True
			'''else:
				otherrundata.append((runnum,PCBEtool,runlength,thedescr))
				allrundata[str(runnum)].append('other')'''
		elif re.search('600',thedescr) and added==False:
			if not re.search('600m',thedescr):
				mr600rundata.append((runnum,PCBEtool,runlength,thedescr))
				allrundata[str(runnum)].append('MR600')
				added=True
		elif added==False:
			del allrundata[str(runnum)]
			'''otherrundata.append((runnum,PCBEtool,runlength,thedescr))
			allrundata[str(runnum)].append('other')'''
		#if runnum==325:
		#	print thedescr.decode('utf-8')
	rowcounter+=1

'''substrates=allrundata.keys()
counter=0
with open('substrate details.csv', 'rb') as f: #substrateID,VendorName,VendorLot,RollNumber,DW Start,DW End,Width,Thickness,Comments
    reader = csv.reader(f)
    for row in reader:
		if counter>0:
			substrate=str(row[0]).lstrip('S00')
			if substrate in substrates:
				allrundata[substrate].append(row[1])
				allrundata[substrate].append(row[2])
				allrundata[substrate].append(row[3])
		elif counter==0:
			counter+=1

porrundata, porMC01runs, porMC02runs=sort_pcbe_tool(porrundata)
mr600rundata, mr600MC01runs, mr600MC02runs=sort_pcbe_tool(mr600rundata)
SPrundata, SPMC01runs, SPMC02runs=sort_pcbe_tool(SPrundata)
otherrundata, otherMC01runs, otherMC02runs=sort_pcbe_tool(otherrundata)

MC01runs=porMC01runs+mr600MC01runs+SPMC01runs+otherMC01runs
MC01runs=sorted(MC01runs,key=lambda x: x[0])
MC02runs=porMC02runs+mr600MC02runs+SPMC02runs+otherMC02runs
MC02runs=sorted(MC02runs,key=lambda x: x[0])

porRuns=porMC01runs+porMC02runs
porRuns=sorted(MC02runs,key=lambda x: x[0])
mr600Runs=mr600MC01runs+mr600MC02runs
mr600Runs=sorted(mr600Runs,key=lambda x: x[0])
SPruns=SPMC01runs+SPMC02runs
SPruns=sorted(SPruns,key=lambda x: x[0])
otherRuns=otherMC01runs+otherMC02runs
otherRuns=sorted(otherRuns,key=lambda x: x[0])'''




	
basepath='Y:\TASK FORCE - Performance drift\list of POR runs'
efffile='all eff data up to 1-16-2015 (402).csv'

#import efficiency data
counter=0
alldata={}
tempeffdata=[]
dtype = [('DateTested', 'datetime64'), ('web_id', 'S10'), ('DW', 'float32'), 
('cell_area', 'float32'), ('substrateID', 'S10'), ('CW', 'float32'), 
('baked', 'int16'), ('mfgtype', 'S10'), ('rejectedreason', 'S10'),
('eff', 'float32'), ('voc', 'float32'), ('jsc', 'float32'), 
('ff', 'float32'), ('rs', 'float32'), ('rsh', 'float32')]
'''
todo: remove lines with blank or other for substrate id, DW, CW, eff, etc
daily check of eff data, append to full list of data with new
email out new stats

0=DateTested
1=web id
2=DW
3=cell area
4=substrate ID
5=CW
6=baked
7=mfgtype
8=rejectedreason
9=eff
10=voc
11=jsc
12=ff
13=rs
14=rsh'''
passed=0
notpassed=0
with open('../raw eff data/'+efffile, 'rb') as f:
    reader = csv.reader(f)
    for row in reader:
		if counter>0 and float(row[9])>=effCutoff:
			passed+=1
			tempeffdata.append(row)
		elif counter==0:
			labelrow=row
		else:
			tempeffdata.append(row)
			notpassed+=1
		counter+=1

#print 'passed: ', passed
#print 'not passed : ', notpassed
print 'total yield : ', float(passed)/(float(notpassed+passed)), '%'

		
		
tempeffdata=sorted(tempeffdata, key=lambda x: x[4])
#tempeffdata=np.array(tempeffdata, dtype=dtype)
#tempeffdata=np.sort(tempeffdata, order=['substrateID','DW'])

firstsubstrate=True
for row in tempeffdata:
	substrate=str(row[4])
	if re.search('R',substrate) or re.search('L',substrate) or substrate==335:
		is13inch=True
	else:
		is13inch=False
	if not is13inch and substrate!='' and substrate!='NA' and substrate!='SPECIAL' and row[2]!='' and row[5]!='' and int(substrate.lstrip('S00'))>=290:
		substrate=str(substrate).lstrip('S00')
		if firstsubstrate:
			current_substrate=substrate
			alldata[current_substrate]=[]
			firstsubstrate=False
		else:
			if substrate!=current_substrate:
				current_substrate=substrate
				alldata[current_substrate]=[]
		alldata[current_substrate].append(row)
		
substrate_list=sorted(alldata.keys())



#yields data according to effCutoff (defined above) and 9.1%
substrateYields={}
for substrate in substrate_list:
	tempYielded=[]
	totalCount=0
	passedCountHi=0
	passedCountLo=0
	for row in alldata[substrate]:
		if float(row[9])>=effCutoff:
			passedCountHi+=1
			tempYielded.append(row)
		if float(row[9])>=9.1:
			passedCountLo+=1
		else:
			pass
		totalCount+=1
	substrateYields[substrate]=[round(100*float(passedCountHi)/float(totalCount),1),round(100*float(passedCountLo)/float(totalCount),1)]
	alldata[substrate]=np.array(tempYielded)

normFitstoData={}
linearFitstoData={} #dict of substrates, with linear fit to [eff,voc,jsc,ff,rs,rsh] [0:5]
residuals={}
for substrate in substrate_list:
	print substrate
	currentData=alldata[substrate]
	if currentData.shape[0]>0:
		currentDW=np.array(currentData[:,2],dtype='float64')
		tempDW=[]
		currentEff=[]
		currentVoc=[]
		currentJsc=[]
		currentFF=[]
		currentRs=[]
		currentRsh=[]
		for each in range(len(currentDW)):
			if currentDW[each]>(currentDW[0]+40) and currentDW[each]<(currentDW[-1]-20):
				tempDW.append(currentDW[each])
				currentEff.append(currentData[each,9])
				currentVoc.append(currentData[each,10])
				currentJsc.append(currentData[each,11])
				currentFF.append(currentData[each,12])
				currentRs.append(currentData[each,13])
				currentRsh.append(currentData[each,14])
		if len(tempDW)>10:
			if (max(tempDW)-min(tempDW))>50.0:
				currentDW=np.array(tempDW,dtype='float64')
				currentEff=np.array(currentEff,dtype='float64')
				currentVoc=np.array(currentVoc,dtype='float64')
				currentJsc=np.array(currentJsc,dtype='float64')
				currentFF=np.array(currentFF,dtype='float64')
				currentRs=np.array(currentRs,dtype='float64')
				currentRsh=np.array(currentRsh,dtype='float64')
				maxDW=max(currentDW)
				minDW=min(currentDW)
				effData=currentData[:,9].astype('float64')
				vocData=currentData[:,10].astype('float64')
				jscData=currentData[:,11].astype('float64')
				ffData=currentData[:,12].astype('float64')
				rsData=currentData[:,13].astype('float64')
				rshData=currentData[:,14].astype('float64')
				
				effFit=np.polyfit(currentDW,currentEff,1,full=True)
				effSlope=(np.poly1d(effFit[0])(maxDW)-np.poly1d(effFit[0])(minDW))/(maxDW-minDW)
				
				vocFit=np.polyfit(currentDW,currentVoc,1,full=True)
				vocSlope=(np.poly1d(vocFit[0])(maxDW)-np.poly1d(vocFit[0])(minDW))/(maxDW-minDW)
				
				jscFit=np.polyfit(currentDW,currentJsc,1,full=True)
				jscSlope=(np.poly1d(jscFit[0])(maxDW)-np.poly1d(jscFit[0])(minDW))/(maxDW-minDW)
				
				ffFit=np.polyfit(currentDW,currentFF,1,full=True)
				ffSlope=(np.poly1d(ffFit[0])(maxDW)-np.poly1d(ffFit[0])(minDW))/(maxDW-minDW)
				
				rsFit=np.polyfit(currentDW,currentRs,1,full=True)
				rsSlope=(np.poly1d(rsFit[0])(maxDW)-np.poly1d(rsFit[0])(minDW))/(maxDW-minDW)
				
				rshFit=np.polyfit(currentDW,currentRsh,1,full=True)
				rshSlope=(np.poly1d(rshFit[0])(maxDW)-np.poly1d(rshFit[0])(minDW))/(maxDW-minDW)
				
				linearFitstoData[substrate]=[effSlope,vocSlope,jscSlope,ffSlope,rsSlope,rshSlope]
				normFitstoData[substrate]=[100*effSlope/np.average(currentEff),100*vocSlope/np.average(currentVoc),100*jscSlope/np.average(currentJsc),100*ffSlope/np.average(currentFF),100*rsSlope/np.average(currentRs),100*rshSlope/np.average(currentRsh)]
				residuals[substrate]=[np.sum(effFit[1]),np.sum(vocFit[1]),np.sum(jscFit[1]),np.sum(ffFit[1]),np.sum(rsFit[1]),np.sum(rshFit[1])]

				savedir='../../eff drift fits/organized by substrate/'+str(substrate)+'/'
				if not os.path.exists(savedir): os.makedirs(savedir)
				
				plt.scatter(currentData[:,2],currentData[:,9])
				plt.plot([minDW,maxDW],[np.poly1d(effFit[0])(minDW),np.poly1d(effFit[0])(maxDW)],color='red',linewidth=3)
				ymax=max([np.poly1d(effFit[0])(minDW),np.poly1d(effFit[0])(maxDW)])*1.05
				ymin=min([np.poly1d(effFit[0])(minDW),np.poly1d(effFit[0])(maxDW)])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' eff fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' eff fit.jpg')
				plt.close()
				
				plt.scatter(currentData[:,2],currentData[:,10])
				plt.plot([minDW,maxDW],[np.poly1d(vocFit[0])(minDW),np.poly1d(vocFit[0])(maxDW)],color='red',linewidth=3)
				ymax=max([np.poly1d(vocFit[0])(minDW),np.poly1d(vocFit[0])(maxDW)])*1.05
				ymin=min([np.poly1d(vocFit[0])(minDW),np.poly1d(vocFit[0])(maxDW)])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' voc fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' voc fit.jpg')
				plt.close()
				
				plt.scatter(currentData[:,2],currentData[:,11])
				plt.plot([minDW,maxDW],[np.poly1d(jscFit[0])(minDW),np.poly1d(jscFit[0])(maxDW)],color='red',linewidth=3)
				ymax=max([np.poly1d(jscFit[0])(minDW),np.poly1d(jscFit[0])(maxDW)])*1.05
				ymin=min([np.poly1d(jscFit[0])(minDW),np.poly1d(jscFit[0])(maxDW)])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' jsc fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' jsc fit.jpg')
				plt.close()
				
				plt.scatter(currentData[:,2],currentData[:,12])
				plt.plot([minDW,maxDW],[np.poly1d(ffFit[0])(minDW),np.poly1d(ffFit[0])(maxDW)],color='red',linewidth=3)
				ymax=max([np.poly1d(ffFit[0])(minDW),np.poly1d(ffFit[0])(maxDW)])*1.05
				ymin=min([np.poly1d(ffFit[0])(minDW),np.poly1d(ffFit[0])(maxDW)])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' ff fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' ff fit.jpg')
				plt.close()
				
				plt.scatter(currentData[:,2],currentData[:,13])
				plt.plot([minDW,maxDW],[np.poly1d(rsFit[0])(minDW),np.poly1d(rsFit[0])(maxDW)],color='red',linewidth=3)
				ymax=max([np.poly1d(rsFit[0])(minDW),np.poly1d(rsFit[0])(maxDW)])*1.05
				ymin=min([np.poly1d(rsFit[0])(minDW),np.poly1d(rsFit[0])(maxDW)])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' rs fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' rs fit.jpg')
				plt.close()
				
				plt.scatter(currentData[:,2],currentData[:,14])
				plt.plot([minDW,maxDW],[np.poly1d(rshFit[0])(minDW),np.poly1d(rshFit[0])(maxDW)],color='red',linewidth=3)
				ymax=max([np.poly1d(rshFit[0])(minDW),np.poly1d(rshFit[0])(maxDW)])*1.05
				ymin=min([np.poly1d(rshFit[0])(minDW),np.poly1d(rshFit[0])(maxDW)])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' rsh fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' rsh fit.jpg')
				plt.close()
				
				
				#normalized plots
				ymaxnorm=max([np.poly1d(effFit[0])(minDW),np.poly1d(effFit[0])(maxDW)])
				plt.scatter(currentData[:,2],effData/ymaxnorm)
				
				plt.plot([minDW,maxDW],[np.poly1d(effFit[0])(minDW)/ymaxnorm,np.poly1d(effFit[0])(maxDW)/ymaxnorm],color='red',linewidth=3)
				ymax=max([np.poly1d(effFit[0])(minDW)/ymaxnorm,np.poly1d(effFit[0])(maxDW)/ymaxnorm])*1.05
				ymin=min([np.poly1d(effFit[0])(minDW)/ymaxnorm,np.poly1d(effFit[0])(maxDW)/ymaxnorm])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' norm eff fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' eff fit - norm.jpg')
				plt.close()
				
				
				ymaxnorm=max([np.poly1d(vocFit[0])(minDW),np.poly1d(vocFit[0])(maxDW)])
				plt.scatter(currentData[:,2],vocData/ymaxnorm)
				plt.plot([minDW,maxDW],[np.poly1d(vocFit[0])(minDW)/ymaxnorm,np.poly1d(vocFit[0])(maxDW)/ymaxnorm],color='red',linewidth=3)
				ymax=max([np.poly1d(vocFit[0])(minDW)/ymaxnorm,np.poly1d(vocFit[0])(maxDW)/ymaxnorm])*1.05
				ymin=min([np.poly1d(vocFit[0])(minDW)/ymaxnorm,np.poly1d(vocFit[0])(maxDW)/ymaxnorm])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' norm voc fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' voc fit - norm.jpg')
				plt.close()
				
				ymaxnorm=max([np.poly1d(jscFit[0])(minDW),np.poly1d(jscFit[0])(maxDW)])
				plt.scatter(currentData[:,2],jscData/ymaxnorm)
				plt.plot([minDW,maxDW],[np.poly1d(jscFit[0])(minDW)/ymaxnorm,np.poly1d(jscFit[0])(maxDW)/ymaxnorm],color='red',linewidth=3)
				ymax=max([np.poly1d(jscFit[0])(minDW)/ymaxnorm,np.poly1d(jscFit[0])(maxDW)/ymaxnorm])*1.05
				ymin=min([np.poly1d(jscFit[0])(minDW)/ymaxnorm,np.poly1d(jscFit[0])(maxDW)/ymaxnorm])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' norm jsc fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' jsc fit - norm.jpg')
				plt.close()
				
				ymaxnorm=max([np.poly1d(ffFit[0])(minDW),np.poly1d(ffFit[0])(maxDW)/ymaxnorm])
				plt.scatter(currentData[:,2],ffData/ymaxnorm)
				plt.plot([minDW,maxDW],[np.poly1d(ffFit[0])(minDW)/ymaxnorm,np.poly1d(ffFit[0])(maxDW)/ymaxnorm],color='red',linewidth=3)
				ymax=max([np.poly1d(ffFit[0])(minDW)/ymaxnorm,np.poly1d(ffFit[0])(maxDW)/ymaxnorm])*1.05
				ymin=min([np.poly1d(ffFit[0])(minDW)/ymaxnorm,np.poly1d(ffFit[0])(maxDW)/ymaxnorm])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' norm ff fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' ff fit - norm.jpg')
				plt.close()
				
				ymaxnorm=max([np.poly1d(rsFit[0])(minDW),np.poly1d(rsFit[0])(maxDW)])
				plt.scatter(currentData[:,2],rsData/ymaxnorm)
				plt.plot([minDW,maxDW],[np.poly1d(rsFit[0])(minDW)/ymaxnorm,np.poly1d(rsFit[0])(maxDW)/ymaxnorm],color='red',linewidth=3)
				ymax=max([np.poly1d(rsFit[0])(minDW)/ymaxnorm,np.poly1d(rsFit[0])(maxDW)/ymaxnorm])*1.05
				ymin=min([np.poly1d(rsFit[0])(minDW)/ymaxnorm,np.poly1d(rsFit[0])(maxDW)/ymaxnorm])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' norm rs fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' rs fit - norm.jpg')
				plt.close()
				
				ymaxnorm=max([np.poly1d(rshFit[0])(minDW),np.poly1d(rshFit[0])(maxDW)])
				plt.scatter(currentData[:,2],rshData/ymaxnorm)
				plt.plot([minDW,maxDW],[np.poly1d(rshFit[0])(minDW)/ymaxnorm,np.poly1d(rshFit[0])(maxDW)/ymaxnorm],color='red',linewidth=3)
				ymax=max([np.poly1d(rshFit[0])(minDW)/ymaxnorm,np.poly1d(rshFit[0])(maxDW)/ymaxnorm])*1.05
				ymin=min([np.poly1d(rshFit[0])(minDW)/ymaxnorm,np.poly1d(rshFit[0])(maxDW)/ymaxnorm])*0.95
				plt.ylim([ymin,ymax])
				plt.title(str(substrate)+' norm rsh fit')
				plt.savefig('../../eff drift fits/organized by substrate/'+str(substrate)+'/'+str(substrate)+' rsh fit - norm.jpg')
				plt.close()
		
sortedsubstratelist=sorted(allrundata.keys())

with open('../../eff drift fits/all runs sorted and with drifts.csv', 'wb') as csvfile:
	spamwriter = csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
	spamwriter.writerow(list(labels[:-1])+['runtype','eff slope','voc slope','jsc slope','ff slope','rs slope','rsh slope']+['norm eff slope','norm voc slope','norm jsc slope','norm ff slope','norm rs slope','norm rsh slope','eff residuals','voc residuals','jsc residuals','ff residuals','rs residuals','rsh residuals'])
	for each in sortedsubstratelist:
		try:
			spamwriter.writerow([each] + allrundata[each][:3] + linearFitstoData[each] + normFitstoData[each] + residuals[each])
		except KeyError:
			pass

		
'''with open('../all runs grouped by tool and POR, MR600.csv', 'wb') as csvfile:
	spamwriter = csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
	spamwriter.writerow(list(labels)+['yield with ' + str(effCutoff) + '% cutoff', 'yield with 9.1% cutoff']+['eff slope','voc slope','jsc slope','ff slope','rs slope','rsh slope']+['norm eff slope','norm voc slope','norm jsc slope','norm ff slope'])
	spamwriter.writerow(['MC01 POR runs'])
	for each in porMC01runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['MC02 POR runs'])
	for each in porMC02runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['MC01 MR600 runs'])
	for each in mr600MC01runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['MC02 MR600 runs'])
	for each in mr600MC02runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['MC01 single pass runs'])
	for each in SPMC01runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['MC02 single pass runs'])
	for each in SPMC02runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['MC01 other runs'])
	for each in otherMC01runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['MC02 other runs'])
	for each in otherMC02runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['all MC01 runs'])
	for each in MC01runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['all MC02 runs'])
	for each in MC02runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['all POR runs'])
	for each in porRuns:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['all MR600 runs'])
	for each in mr600Runs:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['all single-pass runs'])
	for each in SPruns:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)
	spamwriter.writerow([''])
	spamwriter.writerow(['all other runs'])
	for each in otherRuns:
		try:
			each=list(each)+substrateYields[str(each[0])]+linearFitstoData[str(each[0])]+normFitstoData[str(each[0])]
		except KeyError:
			pass
		spamwriter.writerow(each)'''

with open('eff sorted by tool and POR, MR600.csv', 'wb') as csvfile:
	spamwriter = csv.writer(csvfile, delimiter=',')
	spamwriter.writerow(labelrow+['PC/BE tool','run length','VendorName', 'VendorLot','RollNumber','run type'])
	for substrate in substrate_list:
		print substrate
		substrate=str(substrate)
		if not isinstance(alldata[substrate],list):
			alldata[substrate]=alldata[substrate].tolist()
		if int(substrate) in [x[0] for x in porMC01runs] or int(substrate) in [x[0] for x in porMC02runs]:
			for datarow in alldata[substrate]:
				spamwriter.writerow(datarow+allrundata[substrate]+['POR'])
		elif int(substrate) in [x[0] for x in mr600MC01runs] or int(substrate) in [x[0] for x in mr600MC02runs]:
			for datarow in alldata[substrate]:
				spamwriter.writerow(datarow+allrundata[substrate]+['MR600'])
		
			

			

